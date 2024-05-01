import openpyxl


inv_file = openpyxl.load_workbook("inventory.xlsx")

#Grab the Sheet1 of the document
product_list = inv_file["Sheet1"]


#Creating a dictionary to list all the cell data
product_per_supplier = {}
total_value_per_supplier = {}
product_und_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    product_num = product_list.cell(product_row,1).value
    inventory_price = product_list.cell(product_row, 5)
    #Calculation No. Of Products Per Supplier
    if supplier_name in product_per_supplier:
        current_no_of_prod =  product_per_supplier.get(supplier_name)
      
        
        # print("Adding New Supplier")
        product_per_supplier[supplier_name] = current_no_of_prod + 1

    else:
        product_per_supplier[supplier_name] = 1

    #Total value of inventory Per Supplier
    if supplier_name in total_value_per_supplier:
        current_total_value = total_value_per_supplier.get(supplier_name)
        total_value_per_supplier[supplier_name] = current_total_value + inventory * price
    else:
        total_value_per_supplier[supplier_name] = inventory * price


    #Calcualte product with inventory under 10
    if inventory < 10:
        product_und_10_inv[int(product_num)] = inventory

    #Logic for Adding total inventory price
    inventory_price.value = inventory * price


print(inventory_price)
print(product_und_10_inv)
print(product_per_supplier)
print(total_value_per_supplier)


inv_file.save("Inventory_with_total_value.xlsx")








